import openpyxl
from openpyxl.workbook import Workbook

meta_projects = {0 : ('project', 'Проект'),
                 1 : ('name', "Задача"),
                 2 : ('position', "Должность"),
                 3 : ('role', "Роль"),
                 4 : ('price', "Стоимость"),
                 5 : ('netprice', "Стоимость нет"),
                 6 : ('internal', "Проект внутренний")}

meta_employees = {0 : ('name', 'ФИО'),
                 1 : ('position', "Должность"),
                 2 : ('role', "Роль"),
                 3 : ('department', "Отдел"),
                 4 : ('base_salary', "Зарплата"),
                 5 : ('bonus', "Премия"),
                 6 : ('internal', "Сотрудник внутренний")}

tasks = [
    {'name': "Проектирование", 'project': "Разработка сайта госуслуг", 'position': "Инженер 3 категории",
     "role": "Архитектор", 'price': 10, 'netprice': 90000, 'internal': 'Внутренний'},
    {'name': "Реализация", 'project': "Разработка сайта госуслуг", 'position': "Инженер 2 категории",
     "role": "Разработчик", 'price': 20000, 'netprice': 45000, 'internal': 'Внутренний'},
    {'name': "Тестирование", 'project': "Разработка сайта госуслуг", 'position': "Инженер 1 категории",
     "role": "Консультант", 'price': 30000, 'netprice': 28000, 'internal': 'Внутренний'}
]

employees = [
    {'name': "Василий Пупкин", 'base_salary': 10000, 'bonus': 1000, 'internal': 'Внутренний',
     'department': 'Направление разработки', 'position': "Инженер 1 категории", 'role': "Консультант"},
    {'name': "Евгений Сидоров", 'base_salary': 20000, 'bonus': 2000, 'internal': 'Внутренний',
     'department': 'Направление разработки', 'position': "Инженер 2 категории", 'role': "Разработчик"},
    {'name': "Макс Покровский", 'base_salary': 30000, 'bonus': 3000, 'internal': 'Внутренний',
     'department': 'Департамент развития', 'position': "Инженер 3 категории", 'role': "Архитектор"},
]

wb = Workbook()
ws = wb.active
headers = []
for colm in range(len(meta_projects)):
    headers.append(meta_projects[colm][1])
ws.append(headers)
for task in tasks:
    result_list = []
    for colm in range(len(meta_projects)):
        result_list.append(task[meta_projects[colm][0]])
    ws.append(result_list)
ws.title = "Projects"

ws = wb.create_sheet()
headers2 = []
for colm in range(len(meta_employees)):
    headers2.append(meta_employees[colm][1])
ws.append(headers2)
for employee in employees:
    result_list = []
    for colm in range(len(meta_employees)):
        result_list.append(employee[meta_employees[colm][0]])
    ws.append(result_list)
ws.title = "Employees"

# for employee in employees:
#     result_list = [employee['name'], employee['position'], employee['role'], employee['department'],
#                    employee['base_salary'], employee['bonus'], employee['internal']]
#     ws.append(result_list)
# ws.title = "Employees"

wb.save("table.xlsx")

wb = openpyxl.open('table.xlsx', read_only=True)
ws = wb.active
print(ws['B2'].value)

print(ws[2][1].value)

for row in range(1, ws.max_row + 1):
    project = ws[row][0].value
    name = ws[row][1].value
    position = ws[row][2].value
    role = ws[row][3].value
    print(row, project, name, position, role)

ws = wb.worksheets[1]


def get_max_row(workbook, sheet_name):
    ws = workbook[sheet_name]
    for row in range(2, ws.max_row + 1):
        cell = ws[row][0]
        if not cell.value:
            return row - 1
    return ws.max_row - 1


task_number = get_max_row(wb, "Projects")
employees_number = get_max_row(wb, "Employees")

print(task_number, employees_number)




