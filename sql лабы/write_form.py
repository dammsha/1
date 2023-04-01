from openpyxl.workbook import Workbook
from read_db_sql import meta_employee, meta_projects, meta_payment, DB_request


class Create_data_form(DB_request):

    def __init__(self):
        super().__init__()
        self.wb = Workbook()
        ws = self.wb.active
        ws2 = self.wb.create_sheet()
        ws3 = self.wb.create_sheet()
        # считывание элементов кортежа
        ws.title = "Сотрудники"
        # ws.append(['ДОБРО ПОЖАЛОВАТЬ В ФОРМУ ВВОДА! ВВЕДИТЕ ДАННЫЕ СОТРУДНИКОВ'])
        header = []
        # цикл по номеру столбца в мета словаре
        for colm in range(len(meta_employee)):
            header.append(meta_employee[colm])
        ws.append(header)
        # row - словарь по строке работников, db_cont[0] - словарь работников
        for row in self.employee_list:
            row_line = []
            for colm in range(len(meta_employee)):
                # colm - номер обрабатываемого столбца
                # meta_employee[colm] - название обрабатываемого столбца
                # row[meta_employee[colm]] - значение ячейки по названию стобца
                row_line.append(row[meta_employee[colm]])
            ws.append(row_line)

        headers2sheet = []
        ws2.title = "Проекты"
        for colm in range(len(meta_projects)):
            headers2sheet.append(meta_projects[colm])
        headers2sheet.append('Скрытый ключ')
        headers2sheet.append('Сумма затрат')
        headers2sheet.append('Сумма остатка')
        ws2.append(headers2sheet)
        counter = 1
        for row in self.project_list:
            counter += 1
            row_line2 = []
            for colm in range(len(meta_projects)):
                row_line2.append(row[meta_projects[colm]])
            row_line2.append(f'=A{counter}&B{counter}&C{counter}')
            row_line2.append(f'=SUMIF(Выплаты!F:F,G{counter},Выплаты!G:G)')
            row_line2.append(f'=E{counter}-H{counter}')
            ws2.append(row_line2)
        ws2.column_dimensions['G'].hidden = True

        headers3sheet = []
        ws3.title = "Выплаты"
        for colm in range(len(meta_payment)):
            headers3sheet.append(meta_payment[colm])
        headers3sheet.append('Скрытый ключ')
        headers3sheet.append('Сумма оплаты сотруднику')
        headers3sheet.append('Сумма остатка по задаче')
        ws3.append(headers3sheet)
        counter = 1
        for row in self.payment_list:
            counter += 1
            row_line3 = []
            for colm in range(len(meta_payment)):
                row_line3.append(row[meta_payment[colm]])
            row_line3.append(f'=C{counter}&D{counter}&E{counter}')
            row_line3.append(f'=VLOOKUP(B{counter},Сотрудники!$A:$F,2,0)+VLOOKUP(B{counter},Сотрудники!$A:$F,3,0)')
            row_line3.append(f'=VLOOKUP(F{counter},Проекты!$G:$I,3,0)')
            ws3.append(row_line3)
        ws3.column_dimensions['F'].hidden = True



    def save_form(self):
        self.wb.save("Форма ввода.xlsx")


if __name__ == '__main__':
    form1 = Create_data_form()
    form1.save_form()
