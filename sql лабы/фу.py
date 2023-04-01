
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

meta_employee = {0: 'ФИО', 1: 'Зарплата', 2: 'Премия', 3: 'Формальная должность', 4: 'Отдел', 5: 'Признаки'}
meta_projects = {0: 'Проект', 1: 'Наименование задачи', 2: 'Должность', 3: 'Стоимость', 4: 'Чистая стоимость',
                 5: 'Признаки'}
meta_payment = {0: 'Дата', 1: 'Сотрудник', 2: 'Проект', 3: 'Наименование задачи', 4: 'Должность'}

class DB_request():

    def __init__(self):
        self.employee_list = []
        self.project_list = []
        self.payment_list = []
        wb = load_workbook('FBD.xlsx')
        ws = wb['Сотрудники']
        if ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                dict_res = {}
                for meta_colm in meta_employee:
                    dict_res[meta_employee[meta_colm]] = ws[row][meta_colm].value
                self.employee_list.append(dict_res)

        ws_p = wb['Проекты']
        if ws_p.max_row > 1:
            for row in range(2, ws_p.max_row + 1):
                dict_res = {}
                for meta_colm in meta_employee:
                    dict_res[meta_projects[meta_colm]] = ws_p[row][meta_colm].value
                self.project_list.append(dict_res)

        ws_pa = wb['Выплаты']
        if ws_pa.max_row > 1:
            for row in range(2, ws_pa.max_row + 1):
                dict_res = {}
                for meta_colm in meta_payment:
                    dict_res[meta_payment[meta_colm]] = ws_pa[row][meta_colm].value
                self.payment_list.append(dict_res)


if __name__ == '__main__':
    request1 = DB_request()
    for dict in (request1.employee_list, request1.project_list, request1.payment_list):
        for line in dict:
            print(line)