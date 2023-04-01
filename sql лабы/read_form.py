from openpyxl import load_workbook
from read_db_sql import meta_payment
from write_db import Create_new_DB


class Form_request():

    def __init__(self):
        wb = load_workbook('Форма ввода.xlsx')
        ws3 = wb["Выплаты"]
        self.payment_list = []
        if ws3.max_row > 1:
            for row in range(2, ws3.max_row + 1):
                dict_res = {}
                for meta_col in meta_payment:
                    dict_res[meta_payment[meta_col]] = ws3[row][meta_col].value
                self.payment_list.append(dict_res)

    def save(self):
        DB1 = Create_new_DB(payment_list=self.payment_list)
        DB1.save_DB()

if __name__ == '__main__':
    a = Form_request()
    a.save()
